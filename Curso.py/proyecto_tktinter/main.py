import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
import re
import os

nombre_archivo = "datos.xlsx"
if os.path.exists(nombre_archivo):
    wb = load_workbook(nombre_archivo)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["nombre", "edad", "email", "telefono"])  # Encabezados de las columnas

# Guardado de datos

# Función única y corregida para guardar los datos
def guardar_datos():
    nombre = entry_nombre.get()
    edad = entry_edad.get()
    telefono = entry_telefono.get()
    email = entry_email.get()
    # Validar que los campos no estén vacíos
    if not nombre or not edad or not email or not telefono:
        messagebox.showwarning("Advertencia", "Por favor, complete todos los campos.")
        return
    # Validar los números como enteros
    try:
        edad_int = int(edad)
        telefono_int = int(telefono)
    except ValueError:
        messagebox.showwarning(title="Advertencia",
                               message="La edad y Teléfono deben ser números.")
        return
    # Validar el formato del email
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        messagebox.showwarning("Advertencia", "Por favor, ingrese un email válido.")
        return
    # Guardar los datos en el Excel
    ws.append([nombre, edad, email, telefono])
    wb.save(nombre_archivo)
    messagebox.showinfo(title="Información", message="Datos guardados con éxito")
    # Limpiar los campos después de guardar
    entry_nombre.delete(0, tk.END)
    entry_edad.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)

 
root = tk.Tk()
root.title("Formulario de Datos")
root.configure(bg="#5d5c6e")
label_style = {"bg": "#5d5c6e", "fg": "white"}
entry_style = {"bg": "#e0e0e0", "fg": "black"}


label_nombre = tk.Label(root, text="Nombre:", **label_style)
label_nombre.grid(row=0, column=0, padx=10, pady=10)
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=0, column=1, padx=10, pady=10)



label_edad = tk.Label(root, text="Edad:", **label_style)
label_edad.grid(row=1, column=0, padx=10, pady=10)
entry_edad = tk.Entry(root, **entry_style)
entry_edad.grid(row=1, column=1, padx=10, pady=10)


label_email = tk.Label(root, text="Email:", **label_style)
label_email.grid(row=2, column=0, padx=10, pady=10)
entry_email = tk.Entry(root, **entry_style)
entry_email.grid(row=2, column=1, padx=10, pady=10)


label_telefono = tk.Label(root, text="Telefono:", **label_style)
label_telefono.grid(row=3, column=0, padx=10, pady=10)
entry_telefono = tk.Entry(root, **entry_style)
entry_telefono.grid(row=3, column=1, padx=10, pady=10)


boton_guardar = tk.Button(root, text="Guardar", command=guardar_datos,
                          bg="#0377c5", fg="white")
boton_guardar.grid(row=5, column=0, columnspan=2, padx=10, pady=20)

root.mainloop()